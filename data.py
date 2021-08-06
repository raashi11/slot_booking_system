from typing import Final
import openpyxl
import psycopg2

# --------------------------Subject data-------------------------------------->

def get_subject_data():
    lst = []
    file = openpyxl.load_workbook("subject.xlsx")
    f = file.active
    col = f.max_column
    row = f.max_row
    cursor.execute('SELECT sub_id FROM subjects order by sub_id desc')
    g_row=cursor.fetchone()
    if g_row != None:
        sub_id = int(g_row[0])+1
    else:
        sub_id = 1
    for j in range(sub_id, row + 1):
        l = []
        for i in range(1, col + 1):
            data = f.cell(row=j, column=i)
            temp = data.value
            l.append(temp)
        lst.append(l)
    return lst
try:
    conn = None
    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor = conn.cursor()
    query = "INSERT INTO subjects (sub_id,subject) VALUES (%s,%s)"
    details = get_subject_data()
    file = openpyxl.load_workbook("subject.xlsx")
    f = file.active
    col = f.max_column
    row = (f.max_row)
    cursor.execute('SELECT sub_id FROM subjects order by sub_id desc')
    g_row = cursor.fetchone()
    if g_row != None:
        sub_id = int(g_row[0]) + 1
    else:
        sub_id = 1
    for k in range(0, len(details)):

        input1 = (sub_id ,details[k][0])
        cursor.execute(query, input1)
        sub_id=sub_id+1
    conn.commit()
except(Exception, psycopg2.DatabaseError) as error:
    pass
finally:
    if conn is not None:
        pass


# --------------------------batch data-------------------------------------->

def get_batch_data():
    lst = []
    file = openpyxl.load_workbook("batch.xlsx")
    f = file.active
    col = f.max_column
    row = f.max_row
    cursor.execute('SELECT b_id FROM batch order by b_id desc')
    g_row=cursor.fetchone()
    if g_row != None: sub_id = int(g_row[0])+1
    else: sub_id = 1
    for j in range(sub_id, row + 1):
        l = []
        for i in range(1, col + 1):
            data = f.cell(row=j, column=i)
            temp = data.value
            l.append(temp)
        lst.append(l)
    return lst
try:
    conn = None
    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor = conn.cursor()
    query = "INSERT INTO batch (b_id,batch) VALUES (%s,%s)"
    details = get_batch_data()
    file = openpyxl.load_workbook("batch.xlsx")
    f = file.active
    col = f.max_column
    row = (f.max_row)
    cursor.execute('SELECT b_id FROM batch order by b_id desc')
    g_row = cursor.fetchone()
    if g_row != None:
        sub_id = int(g_row[0]) + 1
    else:
        sub_id = 1
    for k in range(0, len(details)):
        input1 =(sub_id, details[k][0])
        cursor.execute(query, input1)
        sub_id=sub_id+1
    conn.commit()

except(Exception, psycopg2.DatabaseError) as error:
    pass
finally:
    if conn is not None:
        pass


# --------------------------Student data-------------------------------------->

def get_student_data():
    lst = []
    file = openpyxl.load_workbook("student.xlsx")
    f = file.active
    col = f.max_column
    row = f.max_row
    cursor.execute('SELECT s_id FROM students order by s_id desc')
    g_row=cursor.fetchone()
    if g_row != None: s_id = int(g_row[0])+1
    else: s_id = 1
    for j in range(s_id, row+1):
        l = []
        for i in range(1, col + 1):
            data = f.cell(row=j, column=i)
            temp = data.value
            l.append(temp)
        lst.append(l)
    return lst
try:
    conn = None
    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor = conn.cursor()
    query = "INSERT INTO students (s_id,name1,batch, batch_id, mobile, email) VALUES (%s,%s,%s,%s, %s,%s)"
    details = get_student_data()
    file = openpyxl.load_workbook("student.xlsx")
    f = file.active
    col = f.max_column
    row = (f.max_row)
    cursor.execute('SELECT s_id FROM students order by s_id desc')
    g_row = cursor.fetchone()
    if g_row != None: s_id = int(g_row[0])+1
    else: s_id = 1
    for k in range(0, len(details)):
        absd = details[k][1]
        ans1 = cursor.execute("SELECT b_id FROM batch WHERE batch = %s ",
                              (absd,)
                              )
        ans3 = cursor.fetchone()
        input1 = (s_id,details[k][0], details[k][1], ans3, details[k][2], details[k][3])
        cursor.execute(query, input1)
        s_id=s_id+1
    conn.commit()
except(Exception, psycopg2.DatabaseError) as error:
    pass
finally:
    if conn is not None:
        pass


# --------------------------Teacher data-------------------------------------->

def get_teacher_data():
    lst = []
    file = openpyxl.load_workbook("teacher.xlsx")
    f = file.active
    col = f.max_column
    row = f.max_row
    cursor.execute('SELECT t_id FROM teachers order by t_id desc')
    g_row = cursor.fetchone()
    if g_row != None:
        sub_id = int(g_row[0]) + 1
    else:
        sub_id = 1
    for j in range(sub_id, row + 1):
        l = []
        for i in range(1, col + 1):
            data = f.cell(row=j, column=i)
            temp = data.value
            l.append(temp)
        lst.append(l)
    return lst
try:
    conn = None
    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor = conn.cursor()
    query = "INSERT INTO teachers (t_id,t_name,subjects, sub_id) VALUES (%s,%s, %s,%s)"
    details = get_teacher_data()
    file = openpyxl.load_workbook("teacher.xlsx")
    f = file.active
    col = f.max_column
    row = (f.max_row)
    cursor.execute('SELECT t_id FROM teachers order by t_id desc')
    g_row = cursor.fetchone()
    if g_row != None:
        sub_id = int(g_row[0]) + 1
    else:
        sub_id = 1
    for k in range(0, len(details)):
        absd = details[k][1]
        ans1 = cursor.execute("SELECT sub_id FROM subjects WHERE subject = %s ",
                              (absd,)
                              )
        ans3 = cursor.fetchone()
        input1 = (sub_id,details[k][0], details[k][1], ans3)
        cursor.execute(query, input1)
        sub_id=sub_id+1
    conn.commit()

except(Exception, psycopg2.DatabaseError) as error:
    pass
finally:
    if conn is not None:
        pass

#------------------------------------------------

def nz(value):
    if value == None:
        return 0
    return value
def slots(time):
    s=int(time[0:2])
    e=int(time[3:])
    ans=[]
    for i in range(s,e):
        a=str(i)
        b=str(i+1)
        ans.append(a+"-"+b)
    return ans

def increment_row():
    increment_row.counter += 1
    return increment_row.counter
increment_row.counter = 0

def get_data():
    lst = []
    file = openpyxl.load_workbook("avail.xlsx")
    f = file.active
    col = f.max_column
    row = f.max_row
    for j in range(1,row+1):
        l = []
        for i in range(1,col + 1):
            data = f.cell(row = j, column= i)
            temp = data.value
            l.append(temp)
        lst.append(l)
    return lst

try:
    conn = None
    conn = psycopg2.connect(
    host="localhost",
    database="sevenclasses",
    user="postgres",
    port="5432",
    password="postgres")
    cursor = conn.cursor()
    query = "INSERT INTO available ( t_id, sub, cday , class_from_to) VALUES (%s,%s,%s,%s)"
    query9 = "SELECT t_id,sub,cday,class_from_to FROM available WHERE t_id=%s AND sub=%s AND cday=%s AND class_from_to=%s;   "
    details = get_data()
    print("details: ")
    print(details)
    file = openpyxl.load_workbook("avail.xlsx")
    f = file.active
    col = f.max_column
    row = (f.max_row)
    for k in range(0, len(details)):
        if(nz(details[k][2]) != 0):
            x = details[k][2].split(", ")
            a=x[0]
            b=x[1]
            q=slots(a)
            if not b == '':
                q0 = slots(b)
                for i in q0:
                    data = (details[k][0], details[k][1], 'monday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        print("inside")
                        cursor.execute(query, data)
            for i in q:
                data = (details[k][0], details[k][1], 'monday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)
        if(nz(details[k][3]) != 0):
            x1 = details[k][3].split(", ")
            a1 = x1[0]
            b1 = x1[1]
            q1 = slots(a1)
            if not b1 == '':
                q2 = slots(b1)
                for i in q2:
                    data = (details[k][0], details[k][1], 'tuesday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)

            for i in q1:
                data = (details[k][0], details[k][1], 'tuesday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)


        if (nz(details[k][4]) != 0):
            x2 = details[k][4].split(", ")
            a2 = x2[0]
            b2 = x2[1]
            q3 = slots(a2)

            if not b2 == '':
                q4 = slots(b2)
                for i in q4:
                    data = (details[k][0], details[k][1], 'wednesday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)

            for i in q3:
                data = (details[k][0], details[k][1], 'wednesday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)


        if (nz(details[k][5]) != 0):
            x3 = details[k][5].split(", ")
            a3 = x3[0]
            b3 = x3[1]
            q5 = slots(a3)
            if not b3 == '':
                q6 = slots(b3)
                for i in q6:
                    data = (details[k][0], details[k][1], 'thursday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)
            for i in q5:
                data = (details[k][0], details[k][1], 'thursday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)


        if (nz(details[k][6]) != 0):
            x4 = details[k][6].split(", ")
            a4 = x4[0]
            b4 = x4[1]
            q7 = slots(a4)

            if not b4 == '':
                q8 = slots(b4)
                for i in q8:
                    data = (details[k][0], details[k][1], 'friday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)
            for i in q7:
                data = (details[k][0], details[k][1], 'friday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)



        if (nz(details[k][7]) != 0):
            x5 = details[k][7].split(", ")
            a5 = x5[0]
            b5 = x5[1]
            q9 = slots(a5)

            if not b5 == '':
                q10 = slots(b5)
                for i in q10:
                    data = (details[k][0], details[k][1], 'saturday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)
            for i in q9:
                data = (details[k][0], details[k][1], 'saturday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)


        if (nz(details[k][8]) != 0):
            x6 = details[k][8].split(", ")
            a6 = x6[0]
            b6 = x6[1]
            q11 = slots(a6)

            if not b6 == '':
                q12= slots(b6)
                for i in q12:

                    data = (details[k][0], details[k][1], 'sunday', i)
                    cursor.execute(query9, data)
                    lb = cursor.fetchone()
                    if lb==None:
                        cursor.execute(query, data)
            for i in q11:
                data = (details[k][0], details[k][1], 'sunday', i)
                cursor.execute(query9, data)
                lb = cursor.fetchone()
                if lb==None:
                    cursor.execute(query, data)
    conn.commit()

except(Exception, psycopg2.DatabaseError) as error:
        print(error)
finally:
    if conn is not None:
        pass

    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor=conn.cursor()
    cursor.execute('SELECT DISTINCT * FROM available ORDER BY t_id, class_from_to;')
    avail=cursor.fetchall()
    print("Available Slots :")
    for row in avail:
        print(row)
cursor.execute('select * from "booked"')
print("Booked Slots :")
avail1=cursor.fetchall()
for row in avail1:
    print(row)


es_id = int(input("Enter student's id : "))
et_id = int(input("Enter teachers's id : "))
esub = input("Enter subject name : ")
ecday = input("Enter day : ")
ecft = input("Enter class timing: ")



def booking_slot1(et_id,esub,ecday,ecft,es_id):
    conn = psycopg2.connect(
        host="localhost",
        database="sevenclasses",
        user="postgres",
        port="5432",
        password="postgres")
    cursor = conn.cursor()

    query = """INSERT INTO booked ( s_id, t_id, sub, cday, class_from_to) VALUES (%s,%s,%s,%s,%s)"""

    #############CHECK1###################
    cursor.execute(
        "SELECT sub FROM available WHERE sub = %s ",
        (esub,)
    )
    row_count = cursor.rowcount
    if row_count == 0:
        print("It Does Not Exist!")
        return
    #############CHECK2###################
    cursor.execute(
        "SELECT s_id FROM students WHERE s_id = %s ",
        (es_id,)
    )
    row_count = cursor.rowcount
    if row_count == 0:
        print("It Does Not Exist!")
        return
    #############CHECK3###################
    cursor.execute(
        "SELECT cday FROM available WHERE cday = %s ",
        (ecday,)
    )
    row_count = cursor.rowcount
    if row_count == 0:
        print("It Does Not Exist!")
        return
    #############CHECK4###################
    cursor.execute(
        "SELECT class_from_to FROM available WHERE class_from_to= %s ",
        (ecft,)
    )
    row_count = cursor.rowcount
    if row_count == 0:
        print("It Does Not Exist!")
        return
        #############CHECK5###################
        cursor.execute(
            "SELECT t_id FROM available WHERE t_id = %s ",
            (et_id,)
        )
        row_count = cursor.rowcount
        if row_count == 0:
            print("It Does Not Exist!")
            return
    cursor.execute(
        "DELETE FROM available WHERE t_id=%s AND sub=%s AND cday=%s AND class_from_to=%s",
        (et_id,esub,ecday,ecft,)
    )
    insert = (es_id,et_id,esub,ecday,ecft)
    cursor.execute(query,insert)
    print("Slot booked!")
    conn.commit()
    conn.close()

booking_slot1(et_id,esub,ecday,ecft,es_id)
